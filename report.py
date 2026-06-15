"""
Vietnam Payroll Calculator - report.py
Xuất báo cáo lương ra file Excel với định dạng chuyên nghiệp.

Output:
    Sheet 1 - "Bảng Lương": Tổng hợp lương tất cả nhân viên
    Sheet 2+ - "Payslip_[Tên]": Phiếu lương chi tiết từng người
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime


# ============================================================
# MÀU SẮC & STYLE
# ============================================================
COLOR_HEADER_BG  = "1F4E79"   # Xanh navy đậm
COLOR_HEADER_FG  = "FFFFFF"   # Chữ trắng
COLOR_TOTAL_BG   = "D6E4F0"   # Xanh nhạt
COLOR_ALT_ROW    = "F2F7FB"   # Xanh rất nhạt (dòng lẻ)
COLOR_ACCENT     = "2E75B6"   # Xanh dương nhạt
COLOR_LABEL_BG   = "BDD7EE"

FMT_VND = '#,##0'             # Định dạng tiền VND

def _thin_border():
    thin = Side(style='thin', color='AAAAAA')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _header_style():
    return Font(name='Arial', bold=True, color=COLOR_HEADER_FG, size=10)

def _body_style(bold=False):
    return Font(name='Arial', bold=bold, size=10)


# ============================================================
# SHEET 1: BẢNG LƯƠNG TỔNG HỢP
# ============================================================

def _write_payroll_table(ws, payroll_results: list, month: int, year: int):
    ws.title = "Bảng Lương"

    # --- Tiêu đề ---
    ws.merge_cells('A1:N1')
    ws['A1'] = f"BẢNG LƯƠNG THÁNG {month:02d}/{year}"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color=COLOR_ACCENT)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:N2')
    ws['A2'] = f"(Áp dụng theo quy định BHXH/BHYT/BHTN và Luật Thuế TNCN hiện hành)"
    ws['A2'].font = Font(name='Arial', italic=True, size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # --- Header cột ---
    headers = [
        "Mã NV", "Họ và Tên",
        "Lương Gross", "Phụ Cấp\n(miễn thuế)",
        "BHXH (8%)", "BHYT (1.5%)", "BHTN (1%)", "Tổng BH",
        "Giảm trừ\nbản thân", "Số NPT", "Giảm trừ\nNPT",
        "TNCT",
        "Thuế TNCN",
        "Lương NET"
    ]

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _header_style()
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[header_row].height = 35

    # --- Dữ liệu ---
    for row_idx, r in enumerate(payroll_results, start=header_row + 1):
        is_alt = (row_idx % 2 == 0)
        fill = PatternFill("solid", fgColor=COLOR_ALT_ROW) if is_alt else None

        row_data = [
            r['id'],
            r['name'],
            r['gross_salary'],
            r['allowance'],
            r['bhxh'],
            r['bhyt'],
            r['bhtn'],
            r['total_insurance'],
            r['deduction_personal'],
            r['dependents'],
            r['deduction_dependent'],
            r['taxable_income'],
            r['pit'],
            r['net_salary'],
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _body_style()
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal='center' if col_idx in (1, 10) else 'right',
                                       vertical='center')
            if fill:
                cell.fill = fill
            # Định dạng tiền
            if col_idx >= 3 and col_idx != 10:
                cell.number_format = FMT_VND

    # --- Dòng tổng cộng ---
    total_row = header_row + len(payroll_results) + 1
    ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
    ws.cell(row=total_row, column=2, value=f"{len(payroll_results)} nhân viên")
    ws.cell(row=total_row, column=2).font = _body_style(bold=True)
    ws.cell(row=total_row, column=2).fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)

    sum_cols = {3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H',
                9: 'I', 11: 'K', 12: 'L', 13: 'M', 14: 'N'}
    data_start = header_row + 1
    data_end   = header_row + len(payroll_results)

    for col_idx, col_letter in sum_cols.items():
        cell = ws.cell(row=total_row, column=col_idx,
                       value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
        cell.font = _body_style(bold=True)
        cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
        cell.border = _thin_border()
        cell.number_format = FMT_VND
        cell.alignment = Alignment(horizontal='right', vertical='center')

    # Cột 10 (Số NPT) tổng
    cell_npt = ws.cell(row=total_row, column=10,
                       value=f"=SUM(J{data_start}:J{data_end})")
    cell_npt.font = _body_style(bold=True)
    cell_npt.fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
    cell_npt.alignment = Alignment(horizontal='center')

    # --- Độ rộng cột ---
    col_widths = [8, 22, 14, 14, 12, 12, 12, 14, 14, 8, 14, 14, 14, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Đóng băng hàng header
    ws.freeze_panes = f"A{header_row + 1}"

    # --- Ghi chú cuối trang ---
    note_row = total_row + 2
    notes = [
        f"Ngày xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "BHXH: 8% | BHYT: 1.5% | BHTN: 1% (phần người lao động đóng)",
        "Giảm trừ bản thân: 11,000,000 đồng/tháng | Giảm trừ NPT: 4,400,000 đồng/người/tháng",
        "Thuế TNCN tính theo biểu lũy tiến 7 bậc (Điều 22, Luật Thuế TNCN)",
    ]
    for i, note in enumerate(notes):
        cell = ws.cell(row=note_row + i, column=1, value=note)
        cell.font = Font(name='Arial', italic=True, size=8, color='888888')
        ws.merge_cells(f'A{note_row + i}:N{note_row + i}')


# ============================================================
# SHEET 2+: PHIẾU LƯƠNG TỪNG NHÂN VIÊN
# ============================================================

def _write_payslip(wb: Workbook, r: dict, month: int, year: int):
    safe_name = r['name'].replace(' ', '_')[:20]
    ws = wb.create_sheet(title=f"Payslip_{safe_name}")

    def label_cell(row, col, text, bold=False, bg=None):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name='Arial', bold=bold, size=10)
        c.alignment = Alignment(vertical='center')
        c.border = _thin_border()
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        return c

    def value_cell(row, col, value, fmt=None, bold=False, color=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name='Arial', bold=bold, size=10,
                      color=color if color else "000000")
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = _thin_border()
        if fmt:
            c.number_format = fmt
        return c

    # --- Tiêu đề phiếu ---
    ws.merge_cells('A1:D1')
    ws['A1'] = "PHIẾU LƯƠNG CÁ NHÂN"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color=COLOR_ACCENT)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:D2')
    ws['A2'] = f"Tháng {month:02d}/{year}"
    ws['A2'].font = Font(name='Arial', bold=True, size=11)
    ws['A2'].alignment = Alignment(horizontal='center')

    # --- Thông tin nhân viên ---
    info = [
        ("Mã nhân viên:", r['id']),
        ("Họ và tên:",    r['name']),
        ("Số NPT đăng ký:", r['dependents']),
    ]
    for row_offset, (lbl, val) in enumerate(info, start=4):
        label_cell(row_offset, 1, lbl, bold=True, bg=COLOR_LABEL_BG)
        ws.merge_cells(f'B{row_offset}:D{row_offset}')
        c = ws.cell(row=row_offset, column=2, value=val)
        c.font = _body_style()
        c.border = _thin_border()

    # --- Bảng chi tiết lương ---
    items = [
        ("THU NHẬP", None, None, True),
        ("Lương gross",          r['gross_salary'],        FMT_VND, False),
        ("Phụ cấp miễn thuế",    r['allowance'],           FMT_VND, False),
        ("KHẤU TRỪ BẢO HIỂM", None, None, True),
        ("BHXH (8%)",            -r['bhxh'],               FMT_VND, False),
        ("BHYT (1.5%)",          -r['bhyt'],               FMT_VND, False),
        ("BHTN (1%)",            -r['bhtn'],               FMT_VND, False),
        ("Tổng khấu trừ BH",     -r['total_insurance'],    FMT_VND, False),
        ("TÍNH THUẾ TNCN", None, None, True),
        ("Giảm trừ bản thân",    -r['deduction_personal'], FMT_VND, False),
        ("Giảm trừ NPT",         -r['deduction_dependent'],FMT_VND, False),
        ("Thu nhập chịu thuế",   r['taxable_income'],      FMT_VND, False),
        ("Thuế TNCN phải nộp",   -r['pit'],                FMT_VND, False),
    ]

    start_row = 8
    for row_offset, (label, value, fmt, is_section) in enumerate(items, start=start_row):
        if is_section:
            ws.merge_cells(f'A{row_offset}:D{row_offset}')
            c = ws.cell(row=row_offset, column=1, value=label)
            c.font = Font(name='Arial', bold=True, size=10, color=COLOR_HEADER_FG)
            c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.border = _thin_border()
        else:
            label_cell(row_offset, 1, label)
            ws.merge_cells(f'B{row_offset}:C{row_offset}')
            ws.cell(row=row_offset, column=2).border = _thin_border()
            color = "C00000" if value and value < 0 else None
            value_cell(row_offset, 4, value, fmt=fmt, color=color)

    # --- Dòng lương NET ---
    net_row = start_row + len(items)
    ws.merge_cells(f'A{net_row}:C{net_row}')
    c = ws.cell(row=net_row, column=1, value="LƯƠNG THỰC NHẬN (NET)")
    c.font = Font(name='Arial', bold=True, size=11, color=COLOR_HEADER_FG)
    c.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = _thin_border()
    ws.row_dimensions[net_row].height = 24

    nc = ws.cell(row=net_row, column=4, value=r['net_salary'])
    nc.font = Font(name='Arial', bold=True, size=11, color=COLOR_HEADER_FG)
    nc.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    nc.number_format = FMT_VND
    nc.alignment = Alignment(horizontal='right', vertical='center')
    nc.border = _thin_border()

    # --- Độ rộng cột ---
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18

    for row in ws.iter_rows():
        for cell in row:
            if cell.row >= 4:
                ws.row_dimensions[cell.row].height = 20


# ============================================================
# HÀM CHÍNH: TẠO FILE EXCEL HOÀN CHỈNH
# ============================================================

def generate_excel_report(payroll_results: list, month: int, year: int,
                           output_path: str):
    """
    Tạo file Excel báo cáo lương hoàn chỉnh.

    Args:
        payroll_results: Kết quả từ calculator.calculate_payroll()
        month: Tháng (1-12)
        year: Năm (VD: 2025)
        output_path: Đường dẫn file xuất ra
    """
    wb = Workbook()

    # Sheet 1: Bảng lương tổng hợp
    ws_main = wb.active
    _write_payroll_table(ws_main, payroll_results, month, year)

    # Sheet 2+: Phiếu lương từng nhân viên
    for result in payroll_results:
        _write_payslip(wb, result, month, year)

    wb.save(output_path)
    print(f"✅ Đã xuất báo cáo: {output_path}")
    print(f"   → {len(payroll_results)} nhân viên | {1 + len(payroll_results)} sheets")
